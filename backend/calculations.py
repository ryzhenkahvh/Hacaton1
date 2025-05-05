import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import csv
import os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

data = pd.read_excel('/content/Задание_2_Потребление_электроэнергии_Апрель.xlsx', header=1)

min_consumption = 10
tariff = 0.2
time_step = 0.5 #в часах

max_days = 30
n_std = 1.4
#тут будут устанавливаться не дефолтные значения для min_consumption, tariff
# min_consumption = float(input("Введите минимальное потребление: "))
# tariff = float(input("Введите тариф: "))

data_time = data.iloc[:, :2]
data_days = data.iloc[:, 0]
data_hours = data_time.iloc[:, 1]
by_time_consumption = data.iloc[:, 2:].sum(axis=1)

sum = data.iloc[:, 2:].sum()
total_sum = data.iloc[:, 2:].sum().sum()
sum_price = total_sum * tariff

optimization = data.iloc[:, 2:].mask(data.iloc[:, 2:] < min_consumption, 0)
opt_sum = optimization.sum()
total_opt_sum = optimization.sum().sum()
opt_price = total_opt_sum * tariff

difference = sum - opt_sum
total_diff = total_sum - total_opt_sum
diff_price = total_diff * tariff

opt_nan = optimization.mask(optimization == 0)
opt_mean = opt_nan.mean()
std_deviation = opt_nan.std()
n_std = 1.4
min_peak_height = opt_mean.fillna(0) + n_std * std_deviation.fillna(0)

peaks = optimization.where(
    optimization > min_peak_height[optimization.columns]
)
peaks_with_time = pd.concat([data_time, peaks], axis=1)
# print(peaks_with_time)

#список не работающих
low_consumption = data.iloc[:, 2:].where((data.iloc[:, 2:] > 0) & (data.iloc[:, 2:] < min_consumption))
low_consumption_with_time = pd.concat([
    data.iloc[:, :2],
    low_consumption
], axis=1).dropna(how='all', subset=low_consumption.columns)

#Упадки
# min_dip_height = opt_mean - n_std * std_deviation

# dips = optimization.where(
#     optimization < min_dip_height[optimization.columns]
# )
# dips_with_time = pd.concat([data_time, dips], axis=1)

#перераспределение пиков на не работающие части

# redistributed_turned_on = data.copy()
# for col in redistributed_turned_on.columns[2:]:
#     peak_mask = (peaks[col].notna())
#     low_mask = (low_consumption[col].notna())

#     if peak_mask.any() and low_mask.any():
#         excess_power = (redistributed_turned_on.loc[peak_mask, col] - opt_mean[col]).sum()

#         available_capacity = (opt_mean[col] - redistributed_turned_on.loc[low_mask, col]).sum()
#         redistribution_ratio = min(1, excess_power / available_capacity) if available_capacity > 0 else 0
#         redistributed_turned_on.loc[peak_mask, col] = opt_mean[col] + 0.5 * (redistributed_turned_on.loc[peak_mask, col] - opt_mean[col])
#         redistributed_turned_on.loc[low_mask, col] += (min_consumption - redistributed_turned_on.loc[low_mask, col]) * redistribution_ratio

#Перераспределение пиков на упадки
# peaks_mean = peaks.mean()
# dips_mean = dips.mean()
# correction_factor = opt_mean - dips_mean

# redistributed_dips = optimization.copy()
# for col in optimization.columns[2:]:
#     redistributed_dips.loc[peaks[col].notna(), col] -= 0.5 * (peaks[col] - peaks_mean[col])
#     redistributed_dips.loc[dips[col].notna(), col] += 0.5 * (dips_mean[col] - dips[col])

# redistributed_dips = pd.concat([data_time, redistributed_dips], axis=1)

copy_data = data.copy()
temp_file = "temp_output.xlsx"
copy_data.to_excel(temp_file, index=False, engine='openpyxl')

wb = load_workbook(temp_file)
ws = wb.active

red_fill = PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF60", end_color="FFFF60", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.column > 2:  # Пропускаем первые 2 столбца

            df_row = cell.row - 2  # Excel строка 2 → pandas строка 0
            df_col = cell.column - 1  # Excel столбец 3 → pandas столбец 2

            original_value = data.iloc[df_row, df_col]
            optimized_value = optimization.iloc[df_row, df_col - 2]  # Учитываем что optimization начинается с 3-го столбца

            if optimized_value == 0 and original_value != 0:
                cell.fill = yellow_fill

            peaks_col = cell.column - 3  # Настройте это смещение по необходимости
            if 0 <= peaks_col < len(peaks.columns) and not pd.isna(peaks.iloc[df_row, peaks_col]):
                if peaks.iloc[df_row, peaks_col] == original_value:
                    cell.fill = red_fill

output_file = "analyzed_marked.xlsx"
ws.title = "Помеченные значения"
wb.save(output_file)

with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
    peaks_with_time.to_excel(writer, sheet_name="Пиковые значения", index=False)
    low_consumption_with_time.to_excel(writer, sheet_name="Не работает, но включено", index=False)
    # redistributed_turned_on.to_excel(writer, sheet_name="Перераспределено(вкл)", index=False, float_format="%.2f")
    # redistributed_dips.to_excel(writer, sheet_name="Перераспределено(упадки)", index=False, float_format="%.2f")
os.remove(temp_file)

#ВРЕМЯ
#ИСПОЛЬЗУЙ ВМЕСТО СРЕДНЕГО МЕДИАНУ(заметка себе)
time_sum = data.iloc[:, 2:].sum(axis=1)
time_mean = data.iloc[:, 2:].mask(data.iloc[:, 2:] == 0).mean(axis=1)
time_max = data.iloc[:, 2:].max(axis=1)
second_time_column = data.iloc[:, 1]

time_together = pd.concat([second_time_column, time_sum, time_mean, time_max], axis=1)

grouped_result = time_together.groupby(time_together.iloc[:, 0]).agg({
    time_together.columns[1]: 'sum',
    time_together.columns[2]: 'mean',
    time_together.columns[3]: 'max',
    time_together.columns[0]: 'count'
})
grouped_result.columns = ['Суммарное', 'Среднее', 'Максимальное', 'Повторы']

########################################################################
min_count = time_together.groupby(time_together.iloc[:, 0]).size().min()
grouped = time_together.groupby(time_together.iloc[:, 0])

resized_sum = grouped.apply(
    lambda x: x.iloc[:min_count, 1].sum()
).to_frame('Суммарное')

# resized_mean = grouped.apply(
#     lambda x: x.iloc[:min_count, 1].sum()
# ).to_frame('Суммарное')
#############################################################################

#написать в сообщении время, в которое суммарно было потреблено больше всего энергии
top_peaks_mean = grouped_result.nlargest(5, ('Среднее')).loc[:, ['Среднее']]
top_peaks_sum = resized_sum.nlargest(5, ('Суммарное'))

# print(top_peaks_sum)

date_sum = data.iloc[:, 2:].sum(axis=1)
date_max = data.iloc[:, 2:].max(axis=1)
first_date_column = data.iloc[:, 0]
date_together = pd.concat([first_date_column, date_sum, date_max], axis=1)

days_consumption = date_together.groupby(date_together.iloc[:, 0]).agg({
     date_together.columns[1]: 'sum',
     date_together.columns[2]: 'max',
})

# if min_count > max_days:
#   days_consumption = days_consumption.tail(max_days)
# dates = days_consumption.head(min_count)

days_consumption.columns = ['Суммарное', 'Максимальное']
print(days_consumption)

# hourly_original = balanced_data.groupby('Время')['Потребление'].sum()

# redistributed_hourly = redistributed_turned_on.iloc[:, 2:].sum(axis=1)
# hourly_redistributed = pd.DataFrame({
#     'Время': data_hours.values[:len(redistributed_hourly)],
#     'Потребление': redistributed_hourly.values
# }).groupby('Время')['Потребление'].sum()

# redistributed_hourly_second = redistributed_dips.iloc[:, 2:].sum(axis=1)
# hourly_redistributed_second = pd.DataFrame({
#     'Время': data_hours.values[:len(redistributed_hourly_second)],
#     'Потребление': redistributed_hourly_second.values
# }).groupby('Время')['Потребление'].sum()

#потребление по часам  (на него наложить оптимизацию???)

# plt.figure(figsize=(14, 7))
# plt.plot(resized_sum.index.astype(str), resized_sum['Суммарное'], label="Потребление", marker='o', color='#1f77b4', linewidth=2)


# plt.plot(grouped_result.index.astype(str), grouped_result['Среднее'], label="Среднее", marker='o', color='#1f77b4', linewidth=2)
#######################################################################################################################
# plt.plot(hourly_redistributed.index.astype(str), hourly_redistributed.values,
#          label="После перераспределения(не работает)", marker='s', color='#2ca02c', linewidth=2)

# plt.plot(hourly_redistributed_second.index.astype(str), hourly_redistributed_second.values,
#          label="После перераспределения(упадки)", marker='s', color='#9ca03a', linewidth=2)
#############################################################################################################

# plt.xlabel("Время")
# plt.ylabel("Суммарное потребление (кВт*ч)")
# plt.title("Суммарное потребление по времении")
# plt.xticks(rotation=45)
# # plt.title("Среднее потребление по времении")
# plt.legend()
# plt.grid(True)
# plt.tight_layout()
# plt.savefig('consumption_by_time.png')
# plt.show()

plt.figure(figsize=(14, 7))
# Создаем столбчатую диаграмму вместо линейного графика
bars = plt.bar(resized_sum.index.astype(str), resized_sum['Суммарное'],
               color='#1f77b4', alpha=0.7, label='Потребление')

plt.xlabel("Время", fontsize=12)
plt.ylabel("Суммарное потребление (кВт·ч)", fontsize=12)
plt.title("Суммарное потребление по времени", fontsize=14, pad=20)
plt.xticks(rotation=45)

# Добавляем горизонтальную линию среднего значения
mean_consumption = resized_sum['Суммарное'].mean()
plt.axhline(y=mean_consumption, color='r', linestyle='--',
            label=f'Среднее: {mean_consumption:.1f} кВт·ч')

plt.legend(fontsize=10)
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.ylim(mean_consumption * 0.8, resized_sum['Суммарное'].max() * 1.1)
plt.tight_layout()

plt.savefig('consumption_by_time.png', dpi=300, bbox_inches='tight')
plt.show()

#потребление по дням
# plt.figure(figsize=(12, 6))
# x = grouped_result.index
# y = grouped_result['Суммарное']
# plt.plot(x, y, label = "Потребление", marker='o', linewidth=2)
# plt.xlabel("Даты")
# plt.ylabel("Суммарное потребление (кВт*ч)")
# plt.title("Суммарное потребление по дням")
# plt.xticks(rotation=45)
# plt.legend()
# plt.grid(True)
# plt.tight_layout()
# plt.savefig('consumption_by_days.png')
# plt.show()

plt.figure(figsize=(14, 7))
# Создаем столбчатую диаграмму
bars = plt.bar(days_consumption.index, days_consumption['Суммарное'],
               color='#1f77b4', alpha=0.7, label='Потребление')

for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height,
             f'{height:.1f}',
             ha='center', va='bottom', fontsize=8)

plt.xlabel("Даты", fontsize=12)
plt.ylabel("Суммарное потребление (кВт·ч)", fontsize=12)
plt.title("Суммарное потребление по дням", fontsize=14, pad=20)
plt.xticks(rotation=45)

mean_consumption = days_consumption['Суммарное'].mean()
plt.axhline(y=mean_consumption, color='r', linestyle='--',
            label=f'Среднее: {mean_consumption:.1f} кВт·ч')

plt.legend(fontsize=10)
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()

plt.savefig('consumption_by_days.png', dpi=300, bbox_inches='tight')
plt.show()



with pd.ExcelWriter('time_analysis.xlsx') as writer:
    grouped_result.to_excel(writer, sheet_name='Потребление по временным интервалам', float_format="%.2f")
    days_consumption.to_excel(writer, sheet_name='Потребление по дням', float_format="%.2f")
    # grouped_result.to_excel(writer, sheet_name='Среднее потребление по временным интервалам', float_format="%.2f")
    top_peaks_sum.to_excel(writer, sheet_name='Топ суммарных пиковых интервалов ', float_format="%.2f")
    top_peaks_mean.to_excel(writer, sheet_name='Топ средних пиковых интервалов ', float_format="%.2f")

report_data = [
    ["Дата анализа", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ["Шаг времени (часы)", time_step],
    ["Топ пиковых периодов", ""],
    ["Суммарно", ""],
    *[[f"{time}", f"{value:.2f} кВт*ч"]
      for time, value in zip(top_peaks_sum.index, top_peaks_sum[('Суммарное')])],
    ["Среднее", ""],
    *[[f"{time}", f"{value:.2f} кВт*ч"]
      for time, value in zip(top_peaks_mean.index, top_peaks_mean[('Среднее')])],
    # ["Среднее потребление в пик", f"{top_peaks[('Среднее')].mean():.2f} кВт*ч"],
    # ["Максимальное потребление в пик", f"{top_peaks[('Максимальное')].max():.2f} кВт*ч"]
]

with open('time_peaks_report.csv', 'w', encoding='utf-8-sig') as f:
    writer = csv.writer(f, delimiter=';')
    writer.writerows(report_data)

#эту же информацию можно добавить в сообщение тг, тк здесь перечислено главное
report_lines = [
    ["Дата создания", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ["Временной отрезок", ],
    ["Минимальное потребление (кВт*ч)", min_consumption],
    ["Тариф (byn)", tariff],
    ["Общая сумма потребления (кВт*ч)", total_sum],
    ["Сумма к оплате (byn)", sum_price],
    ["Потребление не работающих приборов (кВт*ч)", total_diff],
    ["Сумма к оплате за не работающие приборы (byn)", diff_price],
    ["Общая сумма потребления после оптимизации (кВт*ч)", total_opt_sum],
    ["Сумма после оптимизации (byn)", opt_price]
]

with open('consumption_report.csv', 'w', encoding='utf-8-sig') as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerows(report_lines)

print(f"Файл consumption_report.csv успешно создан!")
